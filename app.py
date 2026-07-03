import io
import json
import os
import re
from datetime import datetime
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

def clean_and_parse_raw_txt(raw_text):
    try:
        raw_text = raw_text.strip()
        if not raw_text:
            return None
        normalized = raw_text.replace('\\\\"', '"').replace('\\"', '"')
        
        if normalized.startswith('[') and normalized.endswith(']'):
            try:
                outer_list = json.loads(normalized)
                cleaned_objects = []
                for item in outer_list:
                    if isinstance(item, str):
                        cleaned_objects.append(json.loads(item))
                    elif isinstance(item, dict):
                        cleaned_objects.append(item)
                return cleaned_objects
            except Exception:
                pass

        cleaned_objects = []
        brace_count = 0
        start_idx = -1
        for idx, char in enumerate(normalized):
            if char == '{':
                if brace_count == 0:
                    start_idx = idx
                brace_count += 1
            elif char == '}':
                brace_count -= 1
                if brace_count == 0 and start_idx != -1:
                    potential_json = normalized[start_idx:idx+1]
                    try:
                        parsed_dict = json.loads(potential_json.strip())
                        if isinstance(parsed_dict, dict) and ('cino' in parsed_dict or 'reg_no' in parsed_dict):
                            cleaned_objects.append(parsed_dict)
                    except Exception:
                        pass
                    start_idx = -1
                elif brace_count < 0:
                    brace_count = 0
                    start_idx = -1
        return cleaned_objects if cleaned_objects else None
    except Exception as e:
        print(f"Decoder Failure: {e}")
        return None

def get_court_and_type(case, case_no_str):
    """
    Extracts the precise court location group and determines if it's Civil or Criminal.
    """
    # 1. Classify Court Location Group
    court_info = str(case.get('court_name', '')).lower() + " " + str(case.get('court_no_desg_name', '')).lower() + " " + str(case.get('establishment_name', '')).lower()
    
    if "borivali" in court_info or "cmm" in court_info:
        court_group = "BORIVALI COURT"
    elif "dindoshi" in court_info or "sessions" in court_info or "civil court" in court_info:
        court_group = "DINDOSHI COURT"
    else:
        court_group = "OTHER DISTRICT COURTS"

    # 2. Classify Civil vs Criminal
    case_type = case_no_str.lower()
    criminal_keywords = ['spl.case', 'scc', 'rcc', 'crba', 'bail', 'ex', 'summary', 'state', 'cr', 'criminal']
    petparty = str(case.get('petparty_name', '')).lower()
    
    if any(kw in case_type for kw in criminal_keywords) or "state" in petparty or "police" in petparty or "station" in petparty:
        jurisdiction = "CRIMINAL CASES"
    else:
        jurisdiction = "CIVIL CASES"

    return court_group, jurisdiction

@app.route('/')

def home():

    if os.path.exists('index.html'):

        return send_file('index.html')

    return "Error: index.html file not found.", 404

@app.route('/generate-board', methods=['POST'])
def generate_board_api():
    request_data = request.get_json(silent=True, force=True)
    if not request_data or 'file_content' not in request_data or 'date' not in request_data:
        return jsonify({"error": "Payload Error: Missing data inputs."}), 400

    raw_text_data = request_data['file_content']
    target_date = request_data['date']

    raw_cases = clean_and_parse_raw_txt(raw_text_data)
    if raw_cases is None:
        return jsonify({"error": "Failed to decode data."}), 400

    # Fallback filtering logic to pull matching records from multiple establishments
    filtered_cases = [c for c in raw_cases if c.get('date_next_list') == target_date or c.get('date_last_list') == target_date]
    if not filtered_cases:
        return jsonify({"error": f"No case records found matching date context: {target_date}"}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Board"
    ws.sheet_view.showGridLines = True

    try:
        date_obj = datetime.strptime(target_date, "%Y-%m-%d")
        display_date = date_obj.strftime("%d-%m-%Y")
        display_day = date_obj.strftime("%A")
    except ValueError:
        return jsonify({"error": "Invalid date format setup."}), 400

    # Top Document Title Formatting
    ws['A1'] = "Date:"
    ws['B1'] = display_date
    ws['A2'] = "Day:"
    ws['B2'] = display_day
    for r in ['A1', 'B1', 'A2', 'B2']:
        ws[r].font = Font(name="Calibri", size=11, bold=True)

    ws.merge_cells("A3:J3")
    ws['A3'] = "T R Patel & Associates"
    ws['A3'].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 28

    # Explicit layout hierarchy structure to meet your layout requirement
    order_structure = {
        "DINDOSHI COURT": {"CIVIL CASES": [], "CRIMINAL CASES": []},
        "BORIVALI COURT": {"CIVIL CASES": [], "CRIMINAL CASES": []},
        "OTHER DISTRICT COURTS": {"CIVIL CASES": [], "CRIMINAL CASES": []}
    }

    # Sort each incoming case directly into its specific bucket
    for case in filtered_cases:
        case_no_str = f"{case.get('type_name', '')}/{case.get('reg_no', '')}/{case.get('reg_year', '')}"
        court_group, jurisdiction = get_court_and_type(case, case_no_str)
        order_structure[court_group][jurisdiction].append((case, case_no_str))

    headers = ["Sr No", "Case no.", "Stage of Case", "Time", "Crt no.", "Previous date", "P/D/A", "Plaintiff", "Defendant", "Next Date"]
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    
    start_row = 5
    
    # Process through strict order: DINDOSHI first, then BORIVALI
    for court_name in ["DINDOSHI COURT", "BORIVALI COURT", "OTHER DISTRICT COURTS"]:
        court_data = order_structure[court_name]
        
        # Enforce your custom sorting request: Civil matches first, then Criminal
        for jur_name in ["CIVIL CASES", "CRIMINAL CASES"]:
            cases_list = court_data[jur_name]
            if not cases_list:
                continue # Skip rendering entirely if there's no data matching this segment

            # 1. Create section banner (e.g., DINDOSHI COURT - CIVIL CASES)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
            sec_cell = ws.cell(row=start_row, column=1, value=f"{court_name} - {jur_name}")
            sec_cell.font = Font(name="Calibri", size=11, bold=True, color="1F497D")
            sec_cell.fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
            sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[start_row].height = 24
            start_row += 1

            # 2. Sub-Table Headers
            ws.row_dimensions[start_row].height = 22
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=text)
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            start_row += 1

            # 3. Case Row Injection
            for idx, (case, case_no_str) in enumerate(cases_list, start=1):
                raw_prev_date = case.get('date_last_list', '')
                prev_date_str = ""
                if raw_prev_date:
                    try:
                        prev_date_str = datetime.strptime(raw_prev_date.split()[0], "%Y-%m-%d").strftime("%d-%m-%Y")
                    except Exception:
                        prev_date_str = str(raw_prev_date)

                court_no = str(case.get('court_no', '')) or str(case.get('court_code', '')) or str(case.get('room_no', ''))
                if not court_no:
                    desg_text = str(case.get('court_no_desg_name', ''))
                    desg_digits = re.findall(r'\b\d{1,2}\b', desg_text)
                    court_no = "".join(desg_digits) if desg_digits else desg_text

                case_time_str = ""
                for time_key in ['time', 'case_time', 'hearing_time', 'slot']:
                    if case.get(time_key):
                        case_time_str = str(case[time_key])
                        break

                petparty = str(case.get('petparty_name', '')).lower()
                if "state" in petparty or "police" in petparty or "spl.case child prot" in case_no_str.lower():
                    pda_status = "Acc"
                else:
                    pda_status = "App"

                row_values = [
                    idx,
                    case_no_str,
                    str(case.get('purpose_name', '')).upper(),
                    case_time_str,
                    court_no,
                    prev_date_str,
                    pda_status,
                    case.get('petparty_name', ''),
                    case.get('resparty_name', ''),
                    case.get('note', '') or case.get('date_next_list', '')
                ]

                ws.row_dimensions[start_row].height = 36
                for col_idx, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=start_row, column=col_idx, value=val)
                    cell.font = Font(name="Calibri", size=10)
                    cell.border = thin_border
                    
                    if col_idx in [1, 4, 5, 6, 7]:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                start_row += 1
                
            start_row += 1

    # Apply Auto-Fit width limits across column letters
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val and not isinstance(cell_val, str) and ws.cell(row=row, column=col_idx).coordinate in ws.merged_cells:
                continue
            val = str(cell_val or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"Daily_Board_{target_date}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)